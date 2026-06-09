from __future__ import annotations

import csv
import json
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from knowledge_ops.ai_analysis.models import AI_PLAN_COLUMNS, AI_READINESS_COLUMNS, AIFileRecord, ScenarioEstimate


def write_outputs(result: Dict[str, object], out_dir: str | Path) -> None:
    output = Path(out_dir)
    output.mkdir(parents=True, exist_ok=True)
    records: List[AIFileRecord] = result["records"]  # type: ignore[assignment]
    scenarios: List[ScenarioEstimate] = result["scenarios"]  # type: ignore[assignment]
    sample: List[AIFileRecord] = result["sample_plan"]  # type: ignore[assignment]
    readiness = [record_row(record) for record in records]
    eligibility = readiness
    plan_rows = [plan_row(record) for record in records]
    scenario_rows = [scenario_row(item) for item in scenarios]
    service_rows = service_unit_rows(scenarios)
    guard_rows = budget_guard_rows(scenarios)
    sensitive = [record_row(record) for record in records if record.requires_manual_approval]
    skipped = [record_row(record) for record in records if record.eligibility_status.startswith("SKIPPED") or record.eligibility_status in {"NOT_ELIGIBLE", "LOCAL_ONLY_RECOMMENDED"}]
    sample_rows = [plan_row(record) for record in sample]

    write_csv(output / "ai_readiness_inventory.csv", AI_READINESS_COLUMNS, readiness)
    write_csv(output / "cloud_eligibility.csv", AI_READINESS_COLUMNS, eligibility)
    write_csv(output / "pricing_estimate.csv", ["scenario", "total_cost_usd", "status", "reasons"], scenario_rows)
    write_csv(output / "scenario_summary.csv", ["scenario", "total_cost_usd", "status", "reasons"], scenario_rows)
    write_csv(output / "service_units.csv", ["scenario", "service", "units", "cost_usd"], service_rows)
    write_csv(output / "budget_guard_report.csv", ["scenario", "status", "reason"], guard_rows)
    write_csv(output / "ai_analysis_plan.csv", AI_PLAN_COLUMNS, plan_rows)
    write_csv(output / "ai_sample_plan.csv", AI_PLAN_COLUMNS, sample_rows)
    write_csv(output / "sensitive_cloud_review.csv", AI_READINESS_COLUMNS, sensitive)
    write_csv(output / "skipped_for_cloud.csv", AI_READINESS_COLUMNS, skipped)
    write_csv(output / "errors.csv", ["error"], [])
    write_checklist(output / "cloud_setup_checklist.md")
    write_report(output / "ai_pricing_report.md", records, scenarios)
    write_excel(output / "pricing_estimate.xlsx", scenario_rows, service_rows, readiness, plan_rows, sample_rows, sensitive, guard_rows, skipped)


def record_row(record: AIFileRecord) -> Dict[str, object]:
    data = record.__dict__.copy()
    for key in ["eligible_services", "recommended_features"]:
        data[key] = ";".join(data.get(key, []))
    for key in ["estimated_units_by_service", "estimated_cost_by_service", "scenario_inclusion_flags"]:
        data[key] = json.dumps(data.get(key, {}), ensure_ascii=False)
    return {column: data.get(column, "") for column in AI_READINESS_COLUMNS}


def plan_row(record: AIFileRecord) -> Dict[str, object]:
    cost = sum(record.estimated_cost_by_service.values())
    units = sum(record.estimated_units_by_service.values())
    return {
        "file_id": record.file_id,
        "full_path": record.full_path,
        "mime_type": record.mime_type,
        "size": record.size,
        "local_classification": f"{record.object_suggestion} / {record.department_suggestion} / {record.document_type_suggestion}",
        "sensitivity": record.sensitivity_suggestion,
        "duplicate_status": record.duplicate_kind or "",
        "recommended_cloud_service": record.recommended_service,
        "recommended_features": ";".join(record.recommended_features),
        "scenario_cheap": record.scenario_inclusion_flags.get("cheap", False),
        "scenario_balanced": record.scenario_inclusion_flags.get("balanced", False),
        "scenario_deep": record.scenario_inclusion_flags.get("deep", False),
        "estimated_units": units,
        "estimated_cost_usd": round(cost, 4),
        "risk_level": record.cloud_analysis_risk_level,
        "approval_required": record.requires_manual_approval,
        "approval_reason": record.reason if record.requires_manual_approval else "",
        "final_decision": "",
        "approved_by": "",
        "approved_at": "",
        "execution_status": "",
    }


def scenario_row(item: ScenarioEstimate) -> Dict[str, object]:
    return {
        "scenario": item.scenario,
        "total_cost_usd": round(item.total_cost_usd, 4),
        "status": item.status,
        "reasons": "; ".join(item.reasons),
    }


def service_unit_rows(scenarios: List[ScenarioEstimate]) -> List[Dict[str, object]]:
    rows = []
    for scenario in scenarios:
        for service, units in scenario.service_units.items():
            rows.append({"scenario": scenario.scenario, "service": service, "units": round(units, 2), "cost_usd": round(scenario.service_costs.get(service, 0), 4)})
    return rows


def budget_guard_rows(scenarios: List[ScenarioEstimate]) -> List[Dict[str, object]]:
    rows = []
    for scenario in scenarios:
        if scenario.reasons:
            for reason in scenario.reasons:
                rows.append({"scenario": scenario.scenario, "status": scenario.status, "reason": reason})
        else:
            rows.append({"scenario": scenario.scenario, "status": scenario.status, "reason": ""})
    return rows


def write_csv(path: Path, columns: List[str], rows: Iterable[Dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_report(path: Path, records: List[AIFileRecord], scenarios: List[ScenarioEstimate]) -> None:
    service_counts = Counter(service for record in records for service in record.eligible_services)
    lines = [
        "# Отчет по подготовке AI-анализа и оценке стоимости",
        "",
        f"- Файлов рассмотрено: {len(records)}",
        f"- Подходит для Vision: {service_counts.get('vision', 0)}",
        f"- Подходит для Document AI: {service_counts.get('document_ai', 0)}",
        f"- Подходит для Video Intelligence: {service_counts.get('video_intelligence', 0)}",
        f"- Подходит для Speech-to-Text: {service_counts.get('speech_to_text', 0)}",
        f"- Google Sheets пропущено: {sum(1 for record in records if record.eligibility_status == 'SKIPPED_GOOGLE_SHEET')}",
        f"- Точных дублей исключено: {sum(1 for record in records if record.eligibility_status == 'SKIPPED_DUPLICATE')}",
        f"- Чувствительных файлов с required approval: {sum(1 for record in records if record.requires_manual_approval)}",
        "",
        "## Оценка стоимости по сценариям",
    ]
    for scenario in scenarios:
        lines.append(f"- {scenario.scenario}: ${scenario.total_cost_usd:.2f} ({scenario.status})")
    lines.extend(
        [
            "",
            "## Примечания",
            "- Estimate mode не вызывает Cloud AI APIs.",
            "- Цены являются оценками из конфига; перед реальным запуском их нужно проверить по актуальному Google Cloud Pricing.",
            "- Рекомендуемый следующий шаг: разобрать небольшой sample plan, утвердить scope и только затем явно включать будущий sample analysis.",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_checklist(path: Path) -> None:
    path.write_text(
        """# Чек-лист подготовки Google Cloud

- Выбрать Google Cloud project и проверить billing.
- Включить Cloud Vision API, Document AI API, Video Intelligence API, Speech-to-Text API и Cloud Storage API, если нужен staging.
- Проверить IAM service account с минимальными ролями.
- Хранить credentials только в repository/environment secrets, не в git.
- Настроить `GOOGLE_CLOUD_PROJECT`, `GOOGLE_CLOUD_LOCATION`, опционально `GOOGLE_CLOUD_STAGING_BUCKET`.
- Держать `AI_ANALYSIS_ALLOW_CLOUD_CALLS=false` до утвержденного sample run.
- Создать временный GCS bucket только при необходимости; настроить lifecycle cleanup policy.
- Не хранить полный OCR text, transcripts, thumbnails или keyframes без явного approval.
- Настроить budget alerts, audit logs и data retention.
- Требовать ручной approval перед deep run или sensitive upload.
""",
        encoding="utf-8",
    )


def write_excel(path: Path, scenario_rows, service_rows, readiness, plan_rows, sample_rows, sensitive, guard_rows, skipped) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    add_rows(ws, ["Metric", "Value"], [{"Metric": "Scenarios", "Value": len(scenario_rows)}, {"Metric": "Files", "Value": len(readiness)}])
    add_sheet(wb, "Scenario Summary", ["scenario", "total_cost_usd", "status", "reasons"], scenario_rows)
    add_sheet(wb, "Service Units", ["scenario", "service", "units", "cost_usd"], service_rows)
    add_sheet(wb, "Pricing Estimate", ["scenario", "total_cost_usd", "status", "reasons"], scenario_rows)
    add_sheet(wb, "Cloud Eligibility", AI_READINESS_COLUMNS, readiness)
    add_sheet(wb, "AI Analysis Plan", AI_PLAN_COLUMNS, plan_rows)
    add_sheet(wb, "Sample Plan", AI_PLAN_COLUMNS, sample_rows)
    add_sheet(wb, "Sensitive Review", AI_READINESS_COLUMNS, sensitive)
    add_sheet(wb, "Budget Guards", ["scenario", "status", "reason"], guard_rows)
    add_sheet(wb, "Skipped", AI_READINESS_COLUMNS, skipped)
    add_sheet(wb, "Errors", ["error"], [])
    wb.save(path)


def add_sheet(wb: Workbook, title: str, columns: List[str], rows: List[Dict[str, object]]) -> None:
    ws = wb.create_sheet(title=title[:31])
    add_rows(ws, columns, rows)


def add_rows(ws, columns: List[str], rows: List[Dict[str, object]]) -> None:
    ws.append(columns)
    fill = PatternFill("solid", fgColor="E8EEF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for row in rows:
        ws.append([row.get(column, "") for column in columns])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, column in enumerate(columns, start=1):
        max_len = max([len(str(column))] + [len(str(row.get(column, ""))) for row in rows[:200]])
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 42)
