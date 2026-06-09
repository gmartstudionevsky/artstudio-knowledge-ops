from __future__ import annotations

import csv
import hashlib
import html
import io
import re
import warnings
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple
from xml.etree import ElementTree

import yaml

from knowledge_ops.drive_inventory.config import InventoryConfig
from knowledge_ops.drive_inventory.drive_client import DriveInventoryClient
from knowledge_ops.drive_inventory.models import DOCS_MIME, SHEETS_MIME, SLIDES_MIME, DriveInventoryItem
from knowledge_ops.drive_inventory.normalizer import normalize_name


@dataclass(frozen=True)
class ContentRule:
    rule_id: str
    category: str
    regex_patterns: Tuple[str, ...] = ()
    keywords: Tuple[str, ...] = ()
    negative_patterns: Tuple[str, ...] = ()
    weight: int = 1
    target_fields: Dict[str, str] = field(default_factory=dict)
    sensitivity_flags: Tuple[str, ...] = ()
    explanation: str = ""


@dataclass
class RuleMatch:
    rule_id: str
    category: str
    matched_by: str
    match_count: int
    weight: int
    target_fields: Dict[str, str]
    sensitivity_flags: List[str]
    explanation: str


@dataclass
class ContentInspectionResult:
    extracted: bool
    status: str
    text_hash: str = ""
    length: int = 0
    language_guess: str = ""
    matches: List[RuleMatch] = field(default_factory=list)
    regex_matches_count: int = 0
    sensitivity_flags: List[str] = field(default_factory=list)
    error: str = ""


class ContentRuleEngine:
    def __init__(self, rules: Iterable[ContentRule]):
        self.rules = list(rules)

    @classmethod
    def from_file(cls, path: str | Path) -> "ContentRuleEngine":
        config_path = Path(path)
        if not config_path.exists():
            return cls([])
        with config_path.open("r", encoding="utf-8") as fh:
            data = yaml.safe_load(fh) or {}
        rules = []
        for item in data.get("rules", []):
            rules.append(
                ContentRule(
                    rule_id=item["rule_id"],
                    category=item.get("category", "uncategorized"),
                    regex_patterns=tuple(item.get("regex_patterns", [])),
                    keywords=tuple(item.get("keywords", [])),
                    negative_patterns=tuple(item.get("negative_patterns", [])),
                    weight=int(item.get("weight", 1)),
                    target_fields=item.get("target_fields", {}) or {},
                    sensitivity_flags=tuple(item.get("sensitivity_flags", [])),
                    explanation=item.get("explanation", ""),
                )
            )
        return cls(rules)

    def match(self, text: str) -> List[RuleMatch]:
        normalized = normalize_name(text)
        matches: List[RuleMatch] = []
        for rule in self.rules:
            if any(re.search(pattern, normalized, flags=re.IGNORECASE | re.MULTILINE) for pattern in rule.negative_patterns):
                continue
            keyword_count = sum(1 for keyword in rule.keywords if normalize_name(keyword) in normalized)
            regex_count = 0
            for pattern in rule.regex_patterns:
                regex_count += len(re.findall(pattern, text, flags=re.IGNORECASE | re.MULTILINE))
            if keyword_count or regex_count:
                matched_by = "keyword+regex" if keyword_count and regex_count else "regex" if regex_count else "keyword"
                matches.append(
                    RuleMatch(
                        rule_id=rule.rule_id,
                        category=rule.category,
                        matched_by=matched_by,
                        match_count=keyword_count + regex_count,
                        weight=rule.weight,
                        target_fields=dict(rule.target_fields),
                        sensitivity_flags=list(rule.sensitivity_flags),
                        explanation=rule.explanation,
                    )
                )
        return matches


class ContentInspector:
    def __init__(self, client: DriveInventoryClient, config: InventoryConfig, rule_engine: ContentRuleEngine):
        self.client = client
        self.config = config
        self.rule_engine = rule_engine

    def inspect(self, item: DriveInventoryItem, file_obj: Dict[str, Any]) -> ContentInspectionResult:
        item.content_inspection_enabled = self.config.enable_content_inspection
        if not self.config.enable_content_inspection:
            return ContentInspectionResult(False, "disabled")
        if item.mime_type == SHEETS_MIME:
            return ContentInspectionResult(False, "skipped_google_sheet")
        if item.object_kind == "folder":
            return ContentInspectionResult(False, "skipped_folder")
        max_bytes = int(self.config.max_download_size_mb) * 1024 * 1024
        try:
            text, status = self.extract_text(item, file_obj, max_bytes)
            if not text:
                return ContentInspectionResult(False, status)
            limited = text[: self.config.content_char_limit]
            matches = self.rule_engine.match(limited)
            flags = sorted({flag for match in matches for flag in match.sensitivity_flags})
            regex_count = sum(match.match_count for match in matches if "regex" in match.matched_by)
            return ContentInspectionResult(
                extracted=True,
                status=status,
                text_hash=hashlib.sha256(limited.encode("utf-8", errors="ignore")).hexdigest(),
                length=len(limited),
                language_guess=guess_language(limited),
                matches=matches,
                regex_matches_count=regex_count,
                sensitivity_flags=flags,
            )
        except Exception as exc:
            return ContentInspectionResult(False, "extract_error", error=str(exc)[:500])

    def extract_text(self, item: DriveInventoryItem, file_obj: Dict[str, Any], max_bytes: int) -> Tuple[str, str]:
        mime_type = item.mime_type
        ext = item.extension.lower()
        if mime_type in {DOCS_MIME, SLIDES_MIME}:
            return self.client.export_text(file_obj, max_bytes), "extracted_google_export_text"
        if mime_type.startswith("application/vnd.google-apps."):
            return "", "unsupported_google_native"
        if mime_type.startswith("image/"):
            return ("", "ocr_disabled" if not self.config.enable_ocr else "ocr_not_implemented")
        if ext in ARCHIVE_EXTENSIONS:
            return "", "archive_metadata_only"
        if ext in LEGACY_BINARY_EXTENSIONS:
            return "", "unsupported_legacy_binary"
        if not is_supported_download_type(ext, mime_type):
            return "", "unsupported_type"
        data = self.client.download_bytes(file_obj, max_bytes)
        if ext in PLAIN_TEXT_EXTENSIONS or mime_type.startswith("text/"):
            return extract_plainish(data, ext), "extracted_plain_text"
        if ext == "docx":
            return extract_docx(data), "extracted_docx"
        if ext == "pptx":
            return extract_pptx(data, self.config.content_page_limit), "extracted_pptx"
        if ext == "xlsx":
            if not self.config.enable_excel_content_inspection:
                return "", "excel_content_inspection_disabled"
            return extract_xlsx(data, self.config.content_char_limit), "extracted_xlsx"
        if ext == "pdf" or mime_type == "application/pdf":
            return extract_pdf_text(data, self.config.content_page_limit), "extracted_pdf_text"
        return "", "unsupported_type"


def is_supported_download_type(ext: str, mime_type: str) -> bool:
    return (
        ext in PLAIN_TEXT_EXTENSIONS
        or ext in ZIP_TEXT_EXTENSIONS
        or ext in PDF_EXTENSIONS
        or mime_type.startswith("text/")
        or mime_type == "application/pdf"
    )


def apply_content_result(item: DriveInventoryItem, result: ContentInspectionResult) -> None:
    item.content_extracted = result.extracted
    item.content_extract_status = result.status
    item.content_extract_error = result.error
    item.content_length = result.length
    item.content_text_hash = result.text_hash
    item.content_language_guess = result.language_guess
    item.content_rule_matches = ";".join(match.rule_id for match in result.matches)
    item.content_regex_matches_count = result.regex_matches_count
    item.content_sensitivity_flags = ";".join(result.sensitivity_flags)
    apply_content_classification(item, result.matches)


def apply_content_classification(item: DriveInventoryItem, matches: List[RuleMatch]) -> None:
    if not matches:
        return
    ranked_fields: Dict[str, Dict[str, int]] = {}
    explanations = []
    for match in matches:
        explanations.append(f"{match.rule_id}:{match.explanation or match.category}")
        for field, value in match.target_fields.items():
            ranked_fields.setdefault(field, {})
            ranked_fields[field][value] = ranked_fields[field].get(value, 0) + match.weight * max(1, match.match_count)

    def best(field: str) -> str:
        values = ranked_fields.get(field, {})
        if not values:
            return ""
        return sorted(values.items(), key=lambda pair: (-pair[1], pair[0]))[0][0]

    item.content_based_document_type = best("document_type_suggestion")
    item.content_based_department = best("department_suggestion")
    item.content_based_process = best("process_suggestion")
    item.content_based_object = best("object_suggestion")
    item.content_based_audience = best("audience_suggestion")
    item.content_based_confidence = "high" if len(matches) >= 3 else "medium"
    item.content_based_reason = "; ".join(explanations[:8])
    item.content_classification_boost = "content_rules_matched"

    conflicts = []
    for content_attr, metadata_attr, unknown_values in [
        ("content_based_document_type", "document_type_suggestion", {"", "неизвестно"}),
        ("content_based_department", "department_suggestion", {"", "не определено"}),
        ("content_based_object", "object_suggestion", {"", "объект не определён"}),
        ("content_based_process", "process_suggestion", {"", "не определено"}),
        ("content_based_audience", "audience_suggestion", {"", "не определено", "внутреннее использование"}),
    ]:
        content_value = getattr(item, content_attr)
        metadata_value = getattr(item, metadata_attr)
        if not content_value:
            continue
        if metadata_value in unknown_values:
            setattr(item, metadata_attr, content_value)
        elif metadata_value != content_value:
            conflicts.append(f"{metadata_attr}:{metadata_value}!={content_value}")

    if item.content_sensitivity_flags:
        item.action_recommendation = "SENSITIVE_REVIEW_REQUIRED"
        if item.sensitivity_suggestion in {"unknown", "operational", "media", "archive"}:
            item.sensitivity_suggestion = content_sensitivity_to_inventory(item.content_sensitivity_flags)
    elif conflicts and item.action_recommendation != "SENSITIVE_REVIEW_REQUIRED":
        item.action_recommendation = "REVIEW_REQUIRED"
    item.reason = append_reason(item.reason, "content: " + item.content_based_reason)
    if conflicts:
        item.reason = append_reason(item.reason, "content_metadata_conflict: " + "; ".join(conflicts[:4]))


def extract_plainish(data: bytes, ext: str) -> str:
    text = data.decode("utf-8", errors="replace")
    if ext in {"html", "htm"}:
        text = re.sub(r"<script\b[^>]*>.*?</script>", " ", text, flags=re.IGNORECASE | re.DOTALL)
        text = re.sub(r"<style\b[^>]*>.*?</style>", " ", text, flags=re.IGNORECASE | re.DOTALL)
        text = re.sub(r"<[^>]+>", " ", text)
        text = html.unescape(text)
    if ext == "rtf":
        text = re.sub(r"\\'[0-9a-fA-F]{2}", " ", text)
        text = re.sub(r"\\[a-zA-Z]+-?\d* ?", " ", text)
        text = re.sub(r"[{}]", " ", text)
    if ext == "csv":
        try:
            rows = csv.reader(io.StringIO(text))
            text = "\n".join(" ".join(row) for row in rows)
        except Exception:
            pass
    return compact_text(text)


def extract_docx(data: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        parts = [name for name in archive.namelist() if name.startswith("word/") and name.endswith(".xml")]
        return compact_text(" ".join(xml_text(archive.read(name)) for name in parts))


def extract_pptx(data: bytes, slide_limit: int) -> str:
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        slides = sorted(name for name in archive.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml"))
        return compact_text(" ".join(xml_text(archive.read(name)) for name in slides[:slide_limit]))


def extract_xlsx(data: bytes, char_limit: int) -> str:
    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    chunks = []
    try:
        for sheet in workbook.worksheets[:10]:
            for row in sheet.iter_rows(max_row=200, values_only=True):
                chunks.append(" ".join(str(value) for value in row if value is not None))
                if sum(len(chunk) for chunk in chunks) >= char_limit:
                    return compact_text("\n".join(chunks))
    finally:
        workbook.close()
    return compact_text("\n".join(chunks))


def extract_pdf_text(data: bytes, page_limit: int) -> str:
    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(io.BytesIO(data))
        return compact_text("\n".join((page.extract_text() or "") for page in reader.pages[:page_limit]))
    except Exception:
        text = data.decode("latin-1", errors="ignore")
        chunks = re.findall(r"\(([^()]{3,200})\)", text)
        return compact_text(" ".join(chunks[:2000]))


def xml_text(data: bytes) -> str:
    try:
        root = ElementTree.fromstring(data)
    except ElementTree.ParseError:
        return ""
    return " ".join(node.text or "" for node in root.iter() if node.text)


def compact_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def guess_language(text: str) -> str:
    cyrillic = len(re.findall(r"[А-Яа-яЁё]", text))
    latin = len(re.findall(r"[A-Za-z]", text))
    if cyrillic > latin:
        return "ru"
    if latin:
        return "en"
    return "unknown"


def content_sensitivity_to_inventory(flags: str) -> str:
    parts = set(filter(None, flags.split(";")))
    if {"passport", "snils", "personal_data", "phone", "email"} & parts:
        return "personal_data"
    if {"bank_details", "inn", "bik"} & parts:
        return "financial"
    if {"cadastral_number"} & parts:
        return "owner_data"
    return "unknown"


def append_reason(existing: str, addition: str) -> str:
    if not addition:
        return existing
    if not existing:
        return addition
    if addition in existing:
        return existing
    return f"{existing}; {addition}"
PLAIN_TEXT_EXTENSIONS = {"txt", "csv", "html", "htm", "rtf"}
ZIP_TEXT_EXTENSIONS = {"docx", "pptx", "xlsx"}
PDF_EXTENSIONS = {"pdf"}
ARCHIVE_EXTENSIONS = {"zip", "rar", "7z", "tar", "gz"}
LEGACY_BINARY_EXTENSIONS = {"doc", "ppt", "xls"}
