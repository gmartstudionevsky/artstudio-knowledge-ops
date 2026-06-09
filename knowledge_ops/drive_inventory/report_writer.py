from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from knowledge_ops.drive_inventory.models import INVENTORY_COLUMNS, DriveInventoryItem
from knowledge_ops.drive_inventory.path_analyzer import folder_statistics
from knowledge_ops.drive_inventory.scanner import InventoryResult

FOLDER_COLUMNS = ["file_id", "name", "full_path", "depth", "file_count", "subfolder_count", "top_mime_types", "chaos_signals"]
ERROR_COLUMNS = ["file_id", "name", "error"]
ACCESS_COVERAGE_COLUMNS = ["metric", "value", "notes"]
SUMMARY_COLUMNS = ["dimension", "value", "count"]
RULE_SUMMARY_COLUMNS = ["source", "rule_id", "count"]
EXACT_DUPLICATE_GROUP_COLUMNS = ["duplicate_group_id", "canonical_candidate_id", "row_count", "total_size", "sample_paths"]
CONTENT_INSPECTION_COLUMNS = [
    "file_id",
    "name",
    "full_path",
    "mime_type",
    "extension",
    "content_inspection_enabled",
    "content_extracted",
    "content_extract_status",
    "content_extract_error",
    "content_length",
    "content_text_hash",
    "content_language_guess",
    "content_rule_matches",
    "content_regex_matches_count",
    "content_classification_boost",
    "content_sensitivity_flags",
    "content_based_document_type",
    "content_based_department",
    "content_based_process",
    "content_based_object",
    "content_based_audience",
    "content_based_confidence",
    "content_based_reason",
]
CONTENT_RULE_MATCH_COLUMNS = ["file_id", "name", "full_path", "rule_id", "category", "classification_reason"]
CONTENT_SENSITIVITY_COLUMNS = ["file_id", "name", "full_path", "sensitivity_flag", "sensitivity_suggestion", "action_recommendation"]

RU_HEADERS = {
    "file_id": "ID файла",
    "name": "Название",
    "normalized_name": "Нормализованное название",
    "mime_type": "MIME-тип",
    "object_kind": "Тип объекта",
    "extension": "Расширение",
    "size": "Размер",
    "md5_checksum": "MD5 checksum",
    "content_hash": "Хеш содержимого",
    "export_hash": "Хеш экспорта",
    "image_perceptual_hash": "Перцептивный хеш изображения",
    "web_view_link": "Ссылка Drive",
    "created_time": "Дата создания",
    "modified_time": "Дата изменения",
    "viewed_by_me_time": "Дата просмотра service account",
    "owners": "Владельцы",
    "last_modifying_user": "Последний изменивший пользователь",
    "parents": "Родительские папки",
    "full_path": "Полный путь",
    "depth": "Глубина",
    "drive_id": "ID диска",
    "shared_drive_name": "Название shared drive",
    "trashed": "В корзине",
    "starred": "Помечен звездой",
    "shared": "Общий доступ",
    "permissions_summary": "Кратко о правах",
    "is_google_workspace_native": "Google Workspace native",
    "is_google_sheet_skipped": "Google Sheets пропущен",
    "skip_reason": "Причина пропуска",
    "object_suggestion": "Предложенный объект",
    "department_suggestion": "Предложенное подразделение",
    "function_suggestion": "Предложенный функциональный блок",
    "document_family_suggestion": "Семейство документа",
    "document_type_suggestion": "Тип документа",
    "process_suggestion": "Процесс / тема",
    "audience_suggestion": "Аудитория",
    "sensitivity_suggestion": "Чувствительность / риск",
    "retention_suggestion": "Предложение по хранению",
    "classification_status": "Статус классификации",
    "matched_path_rules": "Сработавшие path-правила",
    "matched_filename_rules": "Сработавшие filename-правила",
    "matched_extension_rules": "Сработавшие extension-правила",
    "matched_sensitivity_rules": "Сработавшие sensitivity-правила",
    "path_confidence": "Уверенность по пути",
    "filename_confidence": "Уверенность по имени",
    "extension_confidence": "Уверенность по расширению",
    "combined_confidence": "Итоговая уверенность",
    "conflict_flags": "Конфликты сигналов",
    "lifecycle_status": "Lifecycle статус",
    "cleanup_category": "Cleanup категория",
    "source_origin": "Источник / origin",
    "media_subtype": "Медиа subtype",
    "image_subtype": "Image subtype",
    "video_subtype": "Video subtype",
    "audio_subtype": "Audio subtype",
    "design_source_subtype": "Design/source subtype",
    "cloud_analysis_candidate": "Кандидат на Cloud AI",
    "priority_for_human_review": "Приоритет ручного разбора",
    "duplicate_group_id": "Группа дублей",
    "duplicate_kind": "Тип дубля",
    "canonical_candidate_id": "Кандидат в канон",
    "action_recommendation": "Рекомендация",
    "confidence": "Уверенность",
    "reason": "Причина разметки",
    "human_decision": "Решение человека",
    "final_location": "Итоговое размещение",
    "comment": "Комментарий",
    "file_count": "Количество файлов",
    "subfolder_count": "Количество подпапок",
    "top_mime_types": "Основные MIME-типы",
    "chaos_signals": "Признаки хаоса",
    "content_inspection_enabled": "Content inspection включен",
    "content_extracted": "Текст извлечен",
    "content_extract_status": "Статус извлечения",
    "content_extract_error": "Ошибка извлечения",
    "content_length": "Длина извлеченного текста",
    "content_text_hash": "Хеш извлеченного текста",
    "content_language_guess": "Предполагаемый язык",
    "content_rule_matches": "Сработавшие правила",
    "content_regex_matches_count": "Количество regex-срабатываний",
    "content_classification_boost": "Усиление классификации содержимым",
    "content_sensitivity_flags": "Флаги чувствительности по содержимому",
    "content_based_document_type": "Тип документа по содержимому",
    "content_based_department": "Подразделение по содержимому",
    "content_based_process": "Процесс по содержимому",
    "content_based_object": "Объект по содержимому",
    "content_based_audience": "Аудитория по содержимому",
    "content_based_confidence": "Уверенность по содержимому",
    "content_based_reason": "Причина по содержимому",
    "rule_id": "ID правила",
    "category": "Категория",
    "classification_reason": "Причина классификации",
    "sensitivity_flag": "Флаг чувствительности",
    "error": "Ошибка",
    "metric": "Метрика",
    "value": "Значение",
    "notes": "Примечание",
    "dimension": "Измерение",
    "value": "Значение",
    "count": "Количество",
    "source": "Источник правила",
    "row_count": "Количество строк",
    "total_size": "Суммарный размер",
    "sample_paths": "Примеры путей",
    "current_path": "Текущий путь",
    "suggested_object": "Предложенный объект",
    "suggested_department": "Предложенное подразделение",
    "suggested_document_type": "Предложенный тип документа",
    "suggested_process": "Предложенный процесс",
    "sensitivity": "Чувствительность",
    "preliminary_action": "Предварительное действие",
    "future_target_area": "Будущая укрупненная зона",
    "approved_by": "Кем утверждено",
    "approved_at": "Дата утверждения",
    "execution_status": "Статус исполнения",
}
MIGRATION_COLUMNS = [
    "file_id",
    "current_path",
    "name",
    "suggested_object",
    "suggested_department",
    "suggested_document_type",
    "suggested_process",
    "sensitivity",
    "duplicate_group_id",
    "duplicate_kind",
    "canonical_candidate_id",
    "preliminary_action",
    "classification_status",
    "lifecycle_status",
    "cleanup_category",
    "source_origin",
    "media_subtype",
    "cloud_analysis_candidate",
    "priority_for_human_review",
    "confidence",
    "reason",
    "human_decision",
    "future_target_area",
    "final_location",
    "approved_by",
    "approved_at",
    "execution_status",
]


def write_reports(result: InventoryResult, out_dir: str | Path) -> None:
    output = Path(out_dir)
    output.mkdir(parents=True, exist_ok=True)
    items = [item for item in result.items if not item.is_google_sheet_skipped]
    folders = folder_statistics(result.items)
    exact = [item for item in items if item.duplicate_kind == "exact"]
    version = [item for item in items if item.duplicate_kind == "version_candidate"]
    semantic = [item for item in items if item.duplicate_kind == "semantic_candidate"]
    classification_review = [item for item in items if item.confidence in {"low", "medium"} or item.document_type_suggestion == "неизвестно"]
    sensitivity_review = [item for item in items if item.action_recommendation == "SENSITIVE_REVIEW_REQUIRED"]
    migration_rows = build_migration_plan(items)
    coverage_rows = build_access_coverage(result)
    rule_summary = build_rule_match_summary(result.items)
    object_summary = build_summary_rows("object", (item.object_suggestion for item in items))
    department_summary = build_summary_rows("department", (item.department_suggestion for item in items))
    document_type_summary = build_summary_rows("document_type", (item.document_type_suggestion for item in items))
    sensitivity_summary = build_summary_rows("sensitivity", (item.sensitivity_suggestion for item in items))
    cleanup_candidates = [item for item in items if item.cleanup_category not in {"", "keep_review", "unknown_review"}]
    system_trash_candidates = [item for item in items if item.cleanup_category == "system_trash_candidate"]
    exact_duplicate_groups = build_exact_duplicate_group_rows(items)
    media_summary = build_summary_rows("media_subtype", (item.media_subtype or item.image_subtype or item.video_subtype or item.audio_subtype or item.design_source_subtype for item in items if item.media_subtype or item.image_subtype or item.video_subtype or item.audio_subtype or item.design_source_subtype))
    unknown_after_v2 = [item for item in items if item.classification_status in {"UNKNOWN", "NEEDS_REVIEW"} or item.combined_confidence in {"unknown", "needs_review"}]

    write_csv(output / "all_objects.csv", INVENTORY_COLUMNS, (item.to_row() for item in result.items))
    write_localized_csv(output / "all_objects_ru.csv", INVENTORY_COLUMNS, (item.to_row() for item in result.items))
    write_csv(output / "inventory.csv", INVENTORY_COLUMNS, (item.to_row() for item in items))
    write_localized_csv(output / "inventory_ru.csv", INVENTORY_COLUMNS, (item.to_row() for item in items))
    write_csv(output / "folders.csv", FOLDER_COLUMNS, folders)
    write_localized_csv(output / "folders_ru.csv", FOLDER_COLUMNS, folders)
    write_csv(output / "skipped_google_sheets.csv", INVENTORY_COLUMNS, (item.to_row() for item in result.skipped_google_sheets))
    write_localized_csv(output / "skipped_google_sheets_ru.csv", INVENTORY_COLUMNS, (item.to_row() for item in result.skipped_google_sheets))
    write_csv(output / "exact_duplicates.csv", INVENTORY_COLUMNS, (item.to_row() for item in exact))
    write_localized_csv(output / "exact_duplicates_ru.csv", INVENTORY_COLUMNS, (item.to_row() for item in exact))
    write_csv(output / "version_duplicate_candidates.csv", INVENTORY_COLUMNS, (item.to_row() for item in version))
    write_localized_csv(output / "version_duplicate_candidates_ru.csv", INVENTORY_COLUMNS, (item.to_row() for item in version))
    write_csv(output / "semantic_duplicate_candidates.csv", INVENTORY_COLUMNS, (item.to_row() for item in semantic))
    write_localized_csv(output / "semantic_duplicate_candidates_ru.csv", INVENTORY_COLUMNS, (item.to_row() for item in semantic))
    write_csv(output / "classification_review.csv", INVENTORY_COLUMNS, (item.to_row() for item in classification_review))
    write_localized_csv(output / "classification_review_ru.csv", INVENTORY_COLUMNS, (item.to_row() for item in classification_review))
    write_csv(output / "sensitivity_review.csv", INVENTORY_COLUMNS, (item.to_row() for item in sensitivity_review))
    write_localized_csv(output / "sensitivity_review_ru.csv", INVENTORY_COLUMNS, (item.to_row() for item in sensitivity_review))
    write_csv(output / "migration_decision_plan.csv", MIGRATION_COLUMNS, migration_rows)
    write_localized_csv(output / "migration_decision_plan_ru.csv", MIGRATION_COLUMNS, migration_rows)
    write_csv(output / "content_inspection.csv", CONTENT_INSPECTION_COLUMNS, content_inspection_rows(result.items))
    write_localized_csv(output / "content_inspection_ru.csv", CONTENT_INSPECTION_COLUMNS, content_inspection_rows(result.items))
    write_csv(output / "content_rule_matches.csv", CONTENT_RULE_MATCH_COLUMNS, content_rule_match_rows(result.items))
    write_localized_csv(output / "content_rule_matches_ru.csv", CONTENT_RULE_MATCH_COLUMNS, content_rule_match_rows(result.items))
    write_csv(output / "content_sensitivity_flags.csv", CONTENT_SENSITIVITY_COLUMNS, content_sensitivity_rows(result.items))
    write_localized_csv(output / "content_sensitivity_flags_ru.csv", CONTENT_SENSITIVITY_COLUMNS, content_sensitivity_rows(result.items))
    write_csv(output / "errors.csv", ERROR_COLUMNS, result.errors)
    write_localized_csv(output / "errors_ru.csv", ERROR_COLUMNS, result.errors)
    write_csv(output / "access_coverage.csv", ACCESS_COVERAGE_COLUMNS, coverage_rows)
    write_localized_csv(output / "access_coverage_ru.csv", ACCESS_COVERAGE_COLUMNS, coverage_rows)
    write_csv(output / "rule_match_summary.csv", RULE_SUMMARY_COLUMNS, rule_summary)
    write_csv(output / "object_classification_summary.csv", SUMMARY_COLUMNS, object_summary)
    write_csv(output / "department_classification_summary.csv", SUMMARY_COLUMNS, department_summary)
    write_csv(output / "document_type_summary.csv", SUMMARY_COLUMNS, document_type_summary)
    write_csv(output / "sensitivity_summary.csv", SUMMARY_COLUMNS, sensitivity_summary)
    write_csv(output / "cleanup_candidates.csv", INVENTORY_COLUMNS, (item.to_row() for item in cleanup_candidates))
    write_csv(output / "system_trash_candidates.csv", INVENTORY_COLUMNS, (item.to_row() for item in system_trash_candidates))
    write_csv(output / "exact_duplicate_groups.csv", EXACT_DUPLICATE_GROUP_COLUMNS, exact_duplicate_groups)
    write_csv(output / "media_classification_summary.csv", SUMMARY_COLUMNS, media_summary)
    write_csv(output / "unknown_after_v2.csv", INVENTORY_COLUMNS, (item.to_row() for item in unknown_after_v2))
    write_tree(output / "drive_structure_tree.md", folders)
    write_audit_report(output / "audit_report.md", result, folders, exact, version, semantic, sensitivity_review)
    write_excel(
        output / "inventory.xlsx",
        result,
        items,
        folders,
        exact,
        version,
        semantic,
        classification_review,
        sensitivity_review,
        migration_rows,
        coverage_rows,
        rule_summary,
        object_summary,
        department_summary,
        document_type_summary,
        sensitivity_summary,
        cleanup_candidates,
        system_trash_candidates,
        exact_duplicate_groups,
        media_summary,
        unknown_after_v2,
    )


def write_csv(path: Path, columns: List[str], rows: Iterable[Dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_localized_csv(path: Path, columns: List[str], rows: Iterable[Dict[str, object]]) -> None:
    localized_columns = [RU_HEADERS.get(column, column) for column in columns]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=localized_columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({RU_HEADERS.get(column, column): row.get(column, "") for column in columns})


def build_migration_plan(items: List[DriveInventoryItem]) -> List[Dict[str, str]]:
    rows = []
    for item in items:
        rows.append(
            {
                "file_id": item.file_id,
                "current_path": item.full_path,
                "name": item.name,
                "suggested_object": item.object_suggestion,
                "suggested_department": item.department_suggestion,
                "suggested_document_type": item.document_type_suggestion,
                "suggested_process": item.process_suggestion,
                "sensitivity": item.sensitivity_suggestion,
                "duplicate_group_id": item.duplicate_group_id,
                "duplicate_kind": item.duplicate_kind,
                "canonical_candidate_id": item.canonical_candidate_id,
                "preliminary_action": item.action_recommendation,
                "classification_status": item.classification_status,
                "lifecycle_status": item.lifecycle_status,
                "cleanup_category": item.cleanup_category,
                "source_origin": item.source_origin,
                "media_subtype": item.media_subtype or item.image_subtype or item.video_subtype or item.audio_subtype or item.design_source_subtype,
                "cloud_analysis_candidate": item.cloud_analysis_candidate,
                "priority_for_human_review": item.priority_for_human_review,
                "confidence": item.confidence,
                "reason": item.reason,
                "human_decision": "",
                "future_target_area": future_target_area(item),
                "final_location": "",
                "approved_by": "",
                "approved_at": "",
                "execution_status": "not_planned",
            }
        )
    return rows


def build_access_coverage(result: InventoryResult) -> List[Dict[str, str]]:
    items = result.items
    files = [item for item in items if item.object_kind == "file"]
    folders = [item for item in items if item.object_kind == "folder"]
    skipped_sheets = result.skipped_google_sheets
    unmapped = [item for item in items if item.full_path.startswith("/Unmapped parent ")]
    drive_counts = Counter(item.drive_id or "(my-drive-or-unknown)" for item in items)
    status_counts = Counter(item.content_extract_status for item in items if item.content_inspection_enabled)
    rows = [
        {"metric": "scope", "value": result.scope, "notes": "Requested scan scope."},
        {"metric": "total_listed_objects", "value": str(len(items)), "notes": "All listed objects, including skipped Google Sheets."},
        {"metric": "files_in_inventory_csv", "value": str(len(files)), "notes": "Non-folder, non-Google-Sheets files."},
        {"metric": "folders", "value": str(len(folders)), "notes": "Folders visible to the credentials."},
        {"metric": "skipped_google_sheets", "value": str(len(skipped_sheets)), "notes": "Metadata-only; content/export/hash skipped by policy."},
        {"metric": "unmapped_parent_paths", "value": str(len(unmapped)), "notes": "Parent folder not present in current listing, often due to scope/access/limit."},
        {"metric": "errors", "value": str(len(result.errors)), "notes": "Per-file processing errors captured without stopping the run."},
        {"metric": "limitations", "value": str(len(result.limitations)), "notes": "See audit_report.md for details."},
    ]
    for drive_id, count in drive_counts.most_common():
        rows.append({"metric": "objects_by_drive_id", "value": str(count), "notes": drive_id})
    for status, count in status_counts.most_common():
        rows.append({"metric": "content_extract_status", "value": str(count), "notes": status})
    return rows


def build_summary_rows(dimension: str, values: Iterable[str]) -> List[Dict[str, str]]:
    counter = Counter(value or "(empty)" for value in values)
    return [
        {"dimension": dimension, "value": value, "count": str(count)}
        for value, count in counter.most_common()
    ]


def build_rule_match_summary(items: List[DriveInventoryItem]) -> List[Dict[str, str]]:
    counters: Dict[str, Counter] = {
        "path": Counter(),
        "filename": Counter(),
        "extension": Counter(),
        "sensitivity": Counter(),
    }
    field_map = {
        "path": "matched_path_rules",
        "filename": "matched_filename_rules",
        "extension": "matched_extension_rules",
        "sensitivity": "matched_sensitivity_rules",
    }
    for item in items:
        for source, field_name in field_map.items():
            for raw_rule in filter(None, getattr(item, field_name).split(";")):
                rule_id = raw_rule.split("@", 1)[0]
                counters[source][rule_id] += 1
    rows = []
    for source, counter in counters.items():
        rows.extend({"source": source, "rule_id": rule_id, "count": str(count)} for rule_id, count in counter.most_common())
    return rows


def build_exact_duplicate_group_rows(items: List[DriveInventoryItem]) -> List[Dict[str, str]]:
    rows = []
    groups = sorted({item.duplicate_group_id for item in items if item.duplicate_kind == "exact" and item.duplicate_group_id})
    for group_id in groups:
        group_items = [item for item in items if item.duplicate_group_id == group_id]
        canonical = next((item.canonical_candidate_id for item in group_items if item.canonical_candidate_id), "")
        rows.append(
            {
                "duplicate_group_id": group_id,
                "canonical_candidate_id": canonical,
                "row_count": str(len(group_items)),
                "total_size": str(sum(int(item.size or 0) for item in group_items)),
                "sample_paths": " | ".join(item.full_path for item in group_items[:5]),
            }
        )
    return rows


def content_inspection_rows(items: List[DriveInventoryItem]) -> List[Dict[str, object]]:
    return [{column: getattr(item, column) for column in CONTENT_INSPECTION_COLUMNS} for item in items]


def content_rule_match_rows(items: List[DriveInventoryItem]) -> List[Dict[str, str]]:
    rows = []
    for item in items:
        for rule_id in filter(None, item.content_rule_matches.split(";")):
            rows.append(
                {
                    "file_id": item.file_id,
                    "name": item.name,
                    "full_path": item.full_path,
                    "rule_id": rule_id,
                    "category": rule_id.split("_")[0],
                    "classification_reason": item.content_based_reason,
                }
            )
    return rows


def content_sensitivity_rows(items: List[DriveInventoryItem]) -> List[Dict[str, str]]:
    rows = []
    for item in items:
        for flag in filter(None, item.content_sensitivity_flags.split(";")):
            rows.append(
                {
                    "file_id": item.file_id,
                    "name": item.name,
                    "full_path": item.full_path,
                    "sensitivity_flag": flag,
                    "sensitivity_suggestion": item.sensitivity_suggestion,
                    "action_recommendation": item.action_recommendation,
                }
            )
    return rows


def future_target_area(item: DriveInventoryItem) -> str:
    if item.duplicate_group_id:
        return "Карантин дублей"
    if item.document_family_suggestion in {"стандарты / SOP / инструкции", "standard_sop", "instruction", "checklist"}:
        return "Канон / стандарты / регламенты / SOP"
    if item.document_family_suggestion in {"шаблоны и формы", "template"}:
        return "Шаблоны и бланки"
    if item.sensitivity_suggestion in {"legal_contract", "owner_contract"}:
        return "Юридические / обязательные документы"
    if item.sensitivity_suggestion in {"financial", "accounting"}:
        return "Финансы / бухгалтерия"
    if item.sensitivity_suggestion in {"HR", "employee_data"}:
        return "Кадры"
    if item.department_suggestion == "маркетинг / SMM / бренд":
        return "Маркетинг и медиа"
    if item.document_family_suggestion in {"отчёты / реестры / финансы", "report", "registry", "financial_document", "accounting_document"}:
        return "Отчёты и аналитика"
    if item.object_suggestion not in {"объект не определён", "не объектный / общий документ"}:
        return "Объектные материалы"
    if item.department_suggestion != "не определено":
        return "Подразделение"
    return "Не определено"


def write_tree(path: Path, folders: List[Dict[str, str]]) -> None:
    lines = ["# Дерево текущей структуры Drive", ""]
    for row in folders:
        indent = "  " * max(0, int(row["depth"]))
        lines.append(
            f"{indent}- {row['name']} | файлов={row['file_count']} | подпапок={row['subfolder_count']} | "
            f"глубина={row['depth']} | типы={row['top_mime_types'] or 'n/a'} | сигналы={row['chaos_signals'] or 'нет'}"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_audit_report(
    path: Path,
    result: InventoryResult,
    folders: List[Dict[str, str]],
    exact: List[DriveInventoryItem],
    version: List[DriveInventoryItem],
    semantic: List[DriveInventoryItem],
    sensitivity: List[DriveInventoryItem],
) -> None:
    items = [item for item in result.items if not item.is_google_sheet_skipped]
    files = [item for item in items if item.object_kind == "file"]
    rule_summary = build_rule_match_summary(result.items)
    path_rule_counts = Counter({row["rule_id"]: int(row["count"]) for row in rule_summary if row["source"] == "path"})
    filename_rule_counts = Counter({row["rule_id"]: int(row["count"]) for row in rule_summary if row["source"] == "filename"})
    sensitivity_rule_counts = Counter({row["rule_id"]: int(row["count"]) for row in rule_summary if row["source"] == "sensitivity"})
    lines = [
        "# Отчет по инвентаризации Google Drive",
        "",
        f"- Дата и время запуска: {result.started_at}",
        f"- Область Drive: {result.scope}",
        f"- Режим: {result.mode}",
        f"- Всего объектов в листинге: {len(result.items)}",
        f"- Папок: {len([item for item in result.items if item.object_kind == 'folder'])}",
        f"- Файлов в реестре: {len(files)}",
        f"- Полный реестр всех увиденных объектов: all_objects.csv / all_objects_ru.csv",
        f"- Google Sheets пропущено: {len(result.skipped_google_sheets)}",
        f"- Объектов с частичным путем из-за unmapped parent: {sum(1 for item in result.items if item.full_path.startswith('/Unmapped parent '))}",
        f"- Content inspection попыток: {sum(1 for item in result.items if item.content_inspection_enabled)}",
        f"- Текст успешно извлечен: {sum(1 for item in result.items if item.content_extracted)}",
        f"- Пропущено по типу/политике: {sum(1 for item in result.items if item.content_extract_status.startswith('skipped') or item.content_extract_status in {'unsupported_type', 'unsupported_legacy_binary', 'archive_metadata_only', 'ocr_disabled'})}",
        f"- OCR извлечений: {sum(1 for item in result.items if item.content_extract_status in {'extracted_image_ocr', 'extracted_pdf_ocr'})}",
        f"- OCR недоступен/не дал текста: {sum(1 for item in result.items if item.content_extract_status in {'ocr_unavailable', 'ocr_no_text', 'ocr_failed'})}",
        f"- Пропущено по размеру: {sum(1 for item in result.items if 'file_too_large' in item.content_extract_error or item.content_extract_status == 'skipped_too_large')}",
        f"- Ошибок извлечения текста: {sum(1 for item in result.items if item.content_extract_status == 'extract_error')}",
        f"- V2 unknown после metadata-классификации: {sum(1 for item in files if item.classification_status in {'UNKNOWN', 'NEEDS_REVIEW'})}",
        "",
        "## Распределения",
        counter_section("MIME-типы", Counter(item.mime_type for item in files)),
        counter_section("Расширения", Counter(item.extension or "(нет)" for item in files)),
        counter_section("Объекты", Counter(item.object_suggestion for item in files)),
        counter_section("Подразделения", Counter(item.department_suggestion for item in files)),
        counter_section("Типы документов", Counter(item.document_type_suggestion for item in files)),
        counter_section("Чувствительность", Counter(item.sensitivity_suggestion for item in files)),
        counter_section("Classification status", Counter(item.classification_status for item in files)),
        counter_section("Cleanup category", Counter(item.cleanup_category for item in files)),
        counter_section("Lifecycle status", Counter(item.lifecycle_status for item in files)),
        counter_section("Priority for human review", Counter(item.priority_for_human_review for item in files)),
        counter_section("Media subtype", Counter(item.media_subtype or item.image_subtype or item.video_subtype or item.audio_subtype or item.design_source_subtype for item in files if item.media_subtype or item.image_subtype or item.video_subtype or item.audio_subtype or item.design_source_subtype)),
        counter_section("Типы документов по содержимому", Counter(item.content_based_document_type for item in files if item.content_based_document_type)),
        counter_section("Флаги чувствительности по содержимому", Counter(flag for item in files for flag in item.content_sensitivity_flags.split(";") if flag)),
        counter_section("Статусы извлечения содержимого", Counter(item.content_extract_status for item in result.items if item.content_inspection_enabled)),
        "",
        "## Папки с высокой плотностью файлов",
    ]
    for row in sorted(folders, key=lambda item: int(item["file_count"]), reverse=True)[:20]:
        lines.append(f"- {row['full_path']}: файлов={row['file_count']}, подпапок={row['subfolder_count']}, сигналы={row['chaos_signals'] or 'нет'}")
    lines.extend(
        [
            "",
            "## Очереди дублей и ручного разбора",
            f"- Строк точных дублей: {len(exact)}; групп: {count_groups(exact)}",
            f"- Кандидатов на версионные дубли: {len(version)}; групп: {count_groups(version)}",
            f"- Кандидатов на смысловые дубли: {len(semantic)}; групп: {count_groups(semantic)}",
            f"- Потенциально чувствительных строк: {len(sensitivity)}",
            f"- Неизвестных типов документов: {sum(1 for item in files if item.document_type_suggestion == 'неизвестно')}",
            f"- Конфликтов metadata/content: {sum(1 for item in files if 'content_metadata_conflict' in item.reason)}",
            f"- Конфликтов metadata V2: {sum(1 for item in files if item.conflict_flags)}",
            f"- Тип определен только по содержимому: {sum(1 for item in files if item.content_based_document_type and item.content_based_document_type == item.document_type_suggestion)}",
            "",
            "## Taxonomy V2",
            counter_section("Top path rules", path_rule_counts),
            counter_section("Top filename rules", filename_rule_counts),
            counter_section("Top sensitivity rules", sensitivity_rule_counts),
            f"- Cleanup candidates: {sum(1 for item in files if item.cleanup_category not in {'', 'keep_review', 'unknown_review'})}",
            f"- System trash candidates: {sum(1 for item in files if item.cleanup_category == 'system_trash_candidate')}",
            f"- Exact duplicate groups: {count_groups(exact)}",
            f"- Owner/legal/financial/HR sensitive rows: {sum(1 for item in files if item.sensitivity_suggestion in {'owner_data', 'owner_contract', 'legal_contract', 'financial', 'HR', 'employee_data'})}",
            "",
            "## Ошибки и ограничения",
            f"- Ошибок: {len(result.errors)}",
        ]
    )
    for limitation in result.limitations or ["В контуре инвентаризации нет операций изменения Google Drive."]:
        lines.append(f"- {limitation}")
    lines.extend(
        [
            "",
            "## Рекомендации для следующего этапа",
            "- Сначала разобрать sensitivity_review.csv / sensitivity_review_ru.csv.",
            "- Проверить группы точных дублей вручную; на этом этапе ничего не удалять.",
            "- Использовать migration_decision_plan.csv как таблицу решений, а не план исполнения.",
            "- Полный аудит запускать после проверки ограниченного прогона.",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def counter_section(title: str, counter: Counter) -> str:
    rows = "; ".join(f"{key}={value}" for key, value in counter.most_common(20))
    return f"- {title}: {rows or 'n/a'}"


def count_groups(items: List[DriveInventoryItem]) -> int:
    return len({item.duplicate_group_id for item in items if item.duplicate_group_id})


def write_excel(
    path: Path,
    result: InventoryResult,
    items: List[DriveInventoryItem],
    folders: List[Dict[str, str]],
    exact: List[DriveInventoryItem],
    version: List[DriveInventoryItem],
    semantic: List[DriveInventoryItem],
    classification_review: List[DriveInventoryItem],
    sensitivity_review: List[DriveInventoryItem],
    migration_rows: List[Dict[str, str]],
    coverage_rows: List[Dict[str, str]],
    rule_summary: List[Dict[str, str]],
    object_summary: List[Dict[str, str]],
    department_summary: List[Dict[str, str]],
    document_type_summary: List[Dict[str, str]],
    sensitivity_summary: List[Dict[str, str]],
    cleanup_candidates: List[DriveInventoryItem],
    system_trash_candidates: List[DriveInventoryItem],
    exact_duplicate_groups: List[Dict[str, str]],
    media_summary: List[Dict[str, str]],
    unknown_after_v2: List[DriveInventoryItem],
) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    add_rows(
        summary,
        ["Metric", "Value"],
        [
            {"Metric": "Run started", "Value": result.started_at},
            {"Metric": "Scope", "Value": result.scope},
            {"Metric": "Mode", "Value": result.mode},
            {"Metric": "Total listed", "Value": len(result.items)},
            {"Metric": "Files inventoried", "Value": len(items)},
            {"Metric": "Folders", "Value": len([item for item in result.items if item.object_kind == "folder"])},
            {"Metric": "Skipped Google Sheets", "Value": len(result.skipped_google_sheets)},
            {"Metric": "Errors", "Value": len(result.errors)},
        ],
    )
    add_sheet(wb, "All Objects", INVENTORY_COLUMNS, [item.to_row() for item in result.items])
    add_sheet(wb, "Inventory", INVENTORY_COLUMNS, [item.to_row() for item in items])
    add_sheet(wb, "Classification V2 Summary", SUMMARY_COLUMNS, object_summary + department_summary + document_type_summary + sensitivity_summary)
    add_sheet(wb, "Objects", SUMMARY_COLUMNS, object_summary)
    add_sheet(wb, "Departments", SUMMARY_COLUMNS, department_summary)
    add_sheet(wb, "Document Types", SUMMARY_COLUMNS, document_type_summary)
    add_sheet(wb, "Sensitivity", SUMMARY_COLUMNS, sensitivity_summary)
    add_sheet(wb, "Cleanup Candidates", INVENTORY_COLUMNS, [item.to_row() for item in cleanup_candidates])
    add_sheet(wb, "System Trash", INVENTORY_COLUMNS, [item.to_row() for item in system_trash_candidates])
    add_sheet(wb, "Media", SUMMARY_COLUMNS, media_summary)
    add_sheet(wb, "Unknown After V2", INVENTORY_COLUMNS, [item.to_row() for item in unknown_after_v2])
    add_sheet(wb, "Rule Matches", RULE_SUMMARY_COLUMNS, rule_summary)
    add_sheet(wb, "Exact Duplicate Groups", EXACT_DUPLICATE_GROUP_COLUMNS, exact_duplicate_groups)
    add_sheet(wb, "Folders", FOLDER_COLUMNS, folders)
    add_sheet(wb, "Skipped Google Sheets", INVENTORY_COLUMNS, [item.to_row() for item in result.skipped_google_sheets])
    add_sheet(wb, "Exact Duplicates", INVENTORY_COLUMNS, [item.to_row() for item in exact])
    add_sheet(wb, "Version Candidates", INVENTORY_COLUMNS, [item.to_row() for item in version])
    add_sheet(wb, "Semantic Candidates", INVENTORY_COLUMNS, [item.to_row() for item in semantic])
    add_sheet(wb, "Classification Review", INVENTORY_COLUMNS, [item.to_row() for item in classification_review])
    add_sheet(wb, "Sensitivity Review", INVENTORY_COLUMNS, [item.to_row() for item in sensitivity_review])
    add_sheet(wb, "Migration Decision Plan", MIGRATION_COLUMNS, migration_rows)
    add_sheet(wb, "Content Inspection", CONTENT_INSPECTION_COLUMNS, content_inspection_rows(result.items))
    add_sheet(wb, "Content Rule Matches", CONTENT_RULE_MATCH_COLUMNS, content_rule_match_rows(result.items))
    add_sheet(wb, "Content Sensitivity", CONTENT_SENSITIVITY_COLUMNS, content_sensitivity_rows(result.items))
    add_sheet(wb, "Access Coverage", ACCESS_COVERAGE_COLUMNS, coverage_rows)
    add_sheet(wb, "Errors", ERROR_COLUMNS, result.errors)
    wb.save(path)


def add_sheet(wb: Workbook, title: str, columns: List[str], rows: List[Dict[str, object]]) -> None:
    ws = wb.create_sheet(title=title[:31])
    add_rows(ws, columns, rows)


def add_rows(ws, columns: List[str], rows: List[Dict[str, object]]) -> None:
    ws.append(columns)
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in rows:
        ws.append([row.get(column, "") for column in columns])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, column in enumerate(columns, start=1):
        max_len = max([len(str(column))] + [len(str(row.get(column, ""))) for row in rows[:200]])
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 42)
